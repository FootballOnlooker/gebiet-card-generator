import os
import re
import pandas as pd
import openpyxl
from PIL import Image, ImageDraw, ImageFont


# =========================
# НАСТРОЙКИ
# =========================
EXCEL_FILE = "C:\\sample_data.xlsx"
DATA_SHEET = "Daten"
NB_SHEET = "NichtBesuchen"
EXTRAS_SHEET = "Extras"

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CARD_WIDTH = 732
MARGIN = 6
BORDER = 2

HEADER_HEIGHT = 60
TABLE_HEADER_HEIGHT = 46
TOP_ROW_HEIGHT = 48
TOP_MIN_ROWS = 8

BOTTOM_TITLE_HEIGHT = 46
BOTTOM_ROW_HEIGHT = 48
BOTTOM_MIN_HEIGHT = 190

EXTRA_PADDING_X = 8
EXTRA_PADDING_Y = 6
EXTRA_LINE_HEIGHT = 26

# ---------- Колонки ----------
COL_STREET = 380
COL_HAUS = 120
COL_STIEGE = 100
COL_PARTEIEN = 120

COLOR_BG = (242, 242, 242)
COLOR_GREEN = (204, 220, 120)
COLOR_YELLOW = (242, 236, 191)
COLOR_EXTRA = (255, 242, 0)
COLOR_BLACK = (0, 0, 0)
COLOR_WHITE = (255, 255, 255)
COLOR_RED = (170, 0, 0)


# =========================
# ШРИФТЫ
# =========================
def load_font(size: int, bold: bool = False):
    candidates = [
        "C:/Windows/Fonts/timesbd.ttf" if bold else "C:/Windows/Fonts/times.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSerif-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation2/LiberationSerif-Regular.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


FONT_TITLE = load_font(30, bold=True)
FONT_HEADER = load_font(22, bold=True)

FONT_STREET = load_font(26, bold=True)
FONT_HAUS = load_font(24, bold=True)
FONT_TEXT = load_font(23, bold=False)
FONT_TEXT_BOLD = load_font(24, bold=True)

FONT_TOTAL = load_font(26, bold=True)

FONT_BOTTOM_TITLE = load_font(24, bold=True)
FONT_BOTTOM_TEXT = load_font(30, bold=True)

FONT_EXTRA = load_font(22, bold=True)


# =========================
# ФУНКЦИИ
# =========================
def text_size(draw, text, font):
    box = draw.textbbox((0, 0), str(text), font=font)
    return box[2] - box[0], box[3] - box[1]


def center_text(draw, box, text, font, fill=COLOR_BLACK):
    x1, y1, x2, y2 = box
    tw, th = text_size(draw, str(text), font)
    x = x1 + (x2 - x1 - tw) / 2
    y = y1 + (y2 - y1 - th) / 2 - 1
    draw.text((x, y), str(text), font=font, fill=fill)


def center_text_auto_font(draw, box, text, font_size, min_size=15, bold=True, fill=COLOR_BLACK):
    x1, y1, x2, y2 = box
    text = str(text)

    for size in range(font_size, min_size - 1, -1):
        font = load_font(size, bold=bold)
        tw, th = text_size(draw, text, font)

        if tw <= (x2 - x1 - 6) and th <= (y2 - y1 - 4):
            center_text(draw, box, text, font, fill)
            return

    font = load_font(min_size, bold=bold)
    center_text(draw, box, text, font, fill)


def right_text(draw, x_right, y_center, text, font, fill=COLOR_BLACK, pad=10):
    tw, th = text_size(draw, str(text), font)
    x = x_right - tw - pad
    y = y_center - th / 2
    draw.text((x, y), str(text), font=font, fill=fill)


def safe_str(value) -> str:
    if pd.isna(value):
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).strip()

    if text.lower() == "nan":
        return ""

    try:
        number = float(text.replace(",", "."))
        if number.is_integer() and "." in text.replace(",", "."):
            return str(int(number))
    except Exception:
        pass

    return text


def count_parties(value) -> int:
    text = safe_str(value)

    if not text:
        return 0

    text = text.replace("–", "-").replace("—", "-").replace(" ", "")

    if re.fullmatch(r"\d+-\d+", text):
        start, end = text.split("-")
        return int(end) - int(start) + 1

    try:
        return int(float(text.replace(",", ".")))
    except Exception:
        return 0


def has_sheet(excel_file: str, sheet_name: str) -> bool:
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    return sheet_name in wb.sheetnames


def wrap_text(draw, text, font, max_width):
    words = str(text).split()
    lines = []
    current_line = ""

    for word in words:
        test_line = word if current_line == "" else current_line + " " + word
        tw, _ = text_size(draw, test_line, font)

        if tw <= max_width:
            current_line = test_line
        else:
            if current_line:
                lines.append(current_line)
            current_line = word

    if current_line:
        lines.append(current_line)

    return lines


def read_sheet_with_excel_rows(excel_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name]

    headers = []
    for cell in ws[1]:
        if cell.value is None:
            headers.append(f"Unnamed_{cell.column}")
        else:
            headers.append(str(cell.value).strip())

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        data = {}
        for i, cell in enumerate(row):
            data[headers[i]] = cell.value
        data["__excel_row__"] = row[0].row
        rows.append(data)

    return pd.DataFrame(rows)


def get_merged_info(excel_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name]

    headers_by_col = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers_by_col[cell.column] = str(cell.value).strip()

    merged_info = {}

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds

        if min_row <= 1:
            continue

        if min_col != max_col:
            continue

        col_name = headers_by_col.get(min_col)

        if not col_name:
            continue

        for row_num in range(min_row, max_row + 1):
            merged_info[(row_num, col_name)] = {
                "start": min_row,
                "end": max_row,
                "span": max_row - min_row + 1,
            }

    return merged_info


def get_merge_span(excel_row, col_name, merged_info):
    info = merged_info.get((excel_row, col_name))
    if info is None:
        return 1
    if excel_row == info["start"]:
        return info["span"]
    return 0


def build_nb_line(row) -> str:
    if "Text" in row.index:
        ready = safe_str(row.get("Text", ""))
        if ready:
            return ready

    street = safe_str(row.get("Straßen Name", ""))
    haus = safe_str(row.get("Haus", ""))
    stiege = safe_str(row.get("Stiegen", ""))
    partei = safe_str(row.get("Parteien", ""))
    top = safe_str(row.get("Top", ""))

    first_part = f"{street} {haus}".strip()

    extra_parts = []

    if stiege:
        extra_parts.append(stiege)

    if top:
        extra_parts.append(top)
    elif partei:
        extra_parts.append(partei)

    if extra_parts:
        return first_part + " / " + " / ".join(extra_parts)

    return first_part


# =========================
# ЧТЕНИЕ EXCEL
# =========================
df_all = read_sheet_with_excel_rows(EXCEL_FILE, DATA_SHEET)
df_nb_all = read_sheet_with_excel_rows(EXCEL_FILE, NB_SHEET)

merged_info = get_merged_info(EXCEL_FILE, DATA_SHEET)

if has_sheet(EXCEL_FILE, EXTRAS_SHEET):
    df_extras_all = read_sheet_with_excel_rows(EXCEL_FILE, EXTRAS_SHEET)
else:
    df_extras_all = pd.DataFrame(columns=["Gebiet", "Text"])


# =========================
# ВСЕ GEBIET
# =========================
df_all["Gebiet_int"] = df_all["Gebiet"].apply(count_parties)
df_nb_all["Gebiet_int"] = df_nb_all["Gebiet"].apply(count_parties)

if "Gebiet" in df_extras_all.columns:
    df_extras_all["Gebiet_int"] = df_extras_all["Gebiet"].apply(count_parties)
else:
    df_extras_all["Gebiet_int"] = 0

ALLE_GEBIETE = sorted(df_all["Gebiet_int"].dropna().unique())

print(f"Найдено Gebiet: {len(ALLE_GEBIETE)}")


# =========================
# СОЗДАНИЕ КАРТОЧЕК
# =========================
for GEBIET in ALLE_GEBIETE:

    if GEBIET == 0:
        continue

    print(f"Создаётся Gebiet {GEBIET}...")

    df = df_all[df_all["Gebiet_int"] == GEBIET].copy()
    df_nb = df_nb_all[df_nb_all["Gebiet_int"] == GEBIET].copy()
    df_extras = df_extras_all[df_extras_all["Gebiet_int"] == GEBIET].copy()

    if df.empty:
        continue

    for col in ["Straßen Name", "Haus", "Stiegen", "Parteien", "Parteien_calc"]:
        if col in df.columns:
            df[col] = df[col].apply(safe_str)

    for col in df_nb.columns:
        df_nb[col] = df_nb[col].apply(safe_str)

    for col in df_extras.columns:
        df_extras[col] = df_extras[col].apply(safe_str)

    total_part = 0

    for _, row in df.iterrows():
        if "Parteien_calc" in df.columns and safe_str(row.get("Parteien_calc", "")):
            total_part += count_parties(row.get("Parteien_calc", ""))
        else:
            total_part += count_parties(row.get("Parteien", ""))

    nb_lines = []

    for _, row in df_nb.iterrows():
        line = build_nb_line(row)
        if line:
            nb_lines.append(line)

    extra_lines = []

    if not df_extras.empty and "Text" in df_extras.columns:
        for _, row in df_extras.iterrows():
            txt = safe_str(row.get("Text", ""))
            if txt:
                extra_lines.append(txt)

    show_extra_block = len(extra_lines) > 0

    temp_img = Image.new("RGB", (CARD_WIDTH, 100), COLOR_WHITE)
    temp_draw = ImageDraw.Draw(temp_img)

    wrapped_extra_lines = []

    if show_extra_block:
        for line in extra_lines:
            wrapped = wrap_text(
                draw=temp_draw,
                text=line,
                font=FONT_EXTRA,
                max_width=CARD_WIDTH - MARGIN * 2 - EXTRA_PADDING_X * 2
            )
            wrapped_extra_lines.append(wrapped)

    top_rows_count = max(TOP_MIN_ROWS, len(df) + 2)
    top_table_height = TABLE_HEADER_HEIGHT + top_rows_count * TOP_ROW_HEIGHT

    extra_block_height = 0

    if show_extra_block:
        for wrapped in wrapped_extra_lines:
            extra_block_height += EXTRA_PADDING_Y * 2 + len(wrapped) * EXTRA_LINE_HEIGHT

    bottom_content_height = max(
        BOTTOM_MIN_HEIGHT,
        len(nb_lines) * BOTTOM_ROW_HEIGHT + 20
    )

    card_height = (
        MARGIN * 2
        + HEADER_HEIGHT
        + top_table_height
        + extra_block_height
        + BOTTOM_TITLE_HEIGHT
        + bottom_content_height
    )

    img = Image.new("RGB", (CARD_WIDTH, card_height), COLOR_WHITE)
    draw = ImageDraw.Draw(img)

    draw.rectangle(
        [MARGIN, MARGIN, CARD_WIDTH - MARGIN, card_height - MARGIN],
        outline=COLOR_BLACK,
        width=BORDER
    )

    y = MARGIN

    draw.rectangle(
        [MARGIN, y, CARD_WIDTH - MARGIN, y + HEADER_HEIGHT],
        fill=COLOR_GREEN,
        outline=COLOR_BLACK,
        width=BORDER
    )

    center_text(
        draw,
        (MARGIN, y, CARD_WIDTH - MARGIN, y + HEADER_HEIGHT),
        f"GEBIET Nr. {GEBIET}",
        FONT_TITLE
    )

    y += HEADER_HEIGHT

    x0 = MARGIN
    x1 = x0 + COL_STREET
    x2 = x1 + COL_HAUS
    x3 = x2 + COL_STIEGE
    x4 = x3 + COL_PARTEIEN

    top_table_bottom = y + top_table_height

    draw.rectangle([x3, y, x4, top_table_bottom], fill=COLOR_YELLOW)
    draw.rectangle([x0, y, x4, top_table_bottom], outline=COLOR_BLACK, width=BORDER)

    for xx in [x1, x2, x3]:
        draw.line([xx, y, xx, top_table_bottom], fill=COLOR_BLACK, width=BORDER)

    draw.line(
        [x0, y + TABLE_HEADER_HEIGHT, x4, y + TABLE_HEADER_HEIGHT],
        fill=COLOR_BLACK,
        width=BORDER
    )

    center_text(draw, (x0, y, x1, y + TABLE_HEADER_HEIGHT), "Straßen Name", FONT_HEADER)
    center_text(draw, (x1, y, x2, y + TABLE_HEADER_HEIGHT), "Haus", FONT_HEADER)
    center_text(draw, (x2, y, x3, y + TABLE_HEADER_HEIGHT), "Stiegen", FONT_HEADER)
    center_text(draw, (x3, y, x4, y + TABLE_HEADER_HEIGHT), "Parteien", FONT_HEADER)

    table_start_y = y + TABLE_HEADER_HEIGHT

    for i in range(top_rows_count):
        yy = table_start_y + (i + 1) * TOP_ROW_HEIGHT

        if i < len(df):
            excel_row = int(df.iloc[i]["__excel_row__"])

            skip_stiegen = (
                merged_info.get((excel_row, "Stiegen")) is not None
                and excel_row < merged_info[(excel_row, "Stiegen")]["end"]
            )

            skip_parteien = (
                merged_info.get((excel_row, "Parteien")) is not None
                and excel_row < merged_info[(excel_row, "Parteien")]["end"]
            )

            draw.line([x0, yy, x1, yy], fill=COLOR_BLACK, width=1)
            draw.line([x1, yy, x2, yy], fill=COLOR_BLACK, width=1)

            if not skip_stiegen:
                draw.line([x2, yy, x3, yy], fill=COLOR_BLACK, width=1)

            if not skip_parteien:
                draw.line([x3, yy, x4, yy], fill=COLOR_BLACK, width=1)

        else:
            draw.line([x0, yy, x4, yy], fill=COLOR_BLACK, width=1)

    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = int(row["__excel_row__"])

        row_top = table_start_y + i * TOP_ROW_HEIGHT
        row_bottom = row_top + TOP_ROW_HEIGHT

        street = safe_str(row.get("Straßen Name", ""))
        haus = safe_str(row.get("Haus", ""))
        stiege = safe_str(row.get("Stiegen", ""))
        part = safe_str(row.get("Parteien", ""))

        center_text(draw, (x0, row_top, x1, row_bottom), street, FONT_STREET)

        center_text_auto_font(
            draw,
            (x1, row_top, x2, row_bottom),
            haus,
            font_size=24,
            min_size=15,
            bold=True
        )

        stiege_span = get_merge_span(excel_row, "Stiegen", merged_info)
        parteien_span = get_merge_span(excel_row, "Parteien", merged_info)

        if stiege_span > 0:
            center_text(
                draw,
                (x2, row_top, x3, row_top + stiege_span * TOP_ROW_HEIGHT),
                stiege,
                FONT_TEXT
            )

        if parteien_span > 0:
            right_text(
                draw,
                x4,
                row_top + (parteien_span * TOP_ROW_HEIGHT) / 2,
                part,
                FONT_TEXT_BOLD
            )

    total_row_index = len(df) + 1
    total_row_top = table_start_y + total_row_index * TOP_ROW_HEIGHT
    total_row_bottom = total_row_top + TOP_ROW_HEIGHT
    total_row_center = (total_row_top + total_row_bottom) / 2

    right_text(draw, x4, total_row_center, str(total_part), FONT_TOTAL)

    y = top_table_bottom

    if show_extra_block:
        for wrapped in wrapped_extra_lines:

            block_height = EXTRA_PADDING_Y * 2 + len(wrapped) * EXTRA_LINE_HEIGHT

            draw.rectangle(
                [MARGIN, y, CARD_WIDTH - MARGIN, y + block_height],
                fill=COLOR_EXTRA,
                outline=COLOR_BLACK,
                width=BORDER
            )

            line_y = y + EXTRA_PADDING_Y

            for text_line in wrapped:
                draw.text(
                    (MARGIN + EXTRA_PADDING_X, line_y),
                    text_line,
                    font=FONT_EXTRA,
                    fill=COLOR_BLACK
                )
                line_y += EXTRA_LINE_HEIGHT

            y += block_height

    draw.rectangle(
        [MARGIN, y, CARD_WIDTH - MARGIN, y + BOTTOM_TITLE_HEIGHT],
        fill=COLOR_WHITE,
        outline=COLOR_BLACK,
        width=BORDER
    )

    center_text(
        draw,
        (MARGIN, y, CARD_WIDTH - MARGIN, y + BOTTOM_TITLE_HEIGHT),
        "Bitte NICHT besuchen",
        FONT_BOTTOM_TITLE
    )

    y += BOTTOM_TITLE_HEIGHT

    bottom_block_bottom = y + bottom_content_height

    draw.rectangle(
        [MARGIN, y, CARD_WIDTH - MARGIN, bottom_block_bottom],
        fill=COLOR_BG,
        outline=COLOR_BLACK,
        width=BORDER
    )

    line_y = y + 10

    for line in nb_lines:
        draw.text(
            (MARGIN + 8, line_y),
            line,
            font=FONT_BOTTOM_TEXT,
            fill=COLOR_RED
        )
        line_y += BOTTOM_ROW_HEIGHT

    out_path = os.path.join(OUTPUT_DIR, f"Gebiet_{GEBIET}.png")
    img.save(out_path)

print("Готово. Все карточки сохранены в папку:", OUTPUT_DIR)
